"""
Script de generación de datos de prueba.
Crea 100 estudiantes ficticios con asistencias, voluntariado, seguimientos y rendimiento.
"""

from __future__ import annotations

import random
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import load_workbook

from config import (
    EXCEL_FILE,
    UNIVERSIDADES, CARRERAS, TIPOS_SEGUIMIENTO,
    ESTADOS_ASISTENCIA, ESTADOS_ESTUDIANTE,
)
from services.excel_manager import ExcelManager

random.seed(42)

NOMBRES = [
    "Carlos", "María", "José", "Ana", "Luis", "Laura", "Miguel", "Sofía",
    "Pedro", "Valentina", "Andrés", "Isabella", "Diego", "Camila", "Jorge",
    "Daniela", "Fernando", "Lucía", "Alejandro", "Natalia", "Eduardo",
    "Gabriela", "Roberto", "Paola", "Sergio", "Verónica", "Héctor", "Carla",
    "Ricardo", "Patricia", "Arturo", "Sandra", "Guillermo", "Monica", "Enrique",
    "Claudia", "Alberto", "Rosa", "Víctor", "Elena",
]

APELLIDOS = [
    "García", "Martínez", "López", "Hernández", "González", "Pérez", "Rodríguez",
    "Sánchez", "Ramírez", "Torres", "Flores", "Rivera", "Gómez", "Díaz", "Cruz",
    "Morales", "Reyes", "Gutiérrez", "Vargas", "Mendoza", "Álvarez", "Castillo",
    "Jiménez", "Moreno", "Ruiz", "Aguilar", "Ortiz", "Medina", "Vega", "Ramos",
    "Núñez", "Rojas", "Herrera", "Domínguez", "Suárez", "Molina", "Ríos",
    "Espinoza", "Campos", "Castro",
]

MONITORES = ["Lic. Ana Mejía", "Lic. Carlos Rivas", "Lic. María López", "Lic. Roberto Chávez"]

ACTIVIDADES_VOLUNTARIADO = [
    "Limpieza de parque comunitario",
    "Apoyo en centro de salud",
    "Tutoría a estudiantes de secundaria",
    "Recolección de alimentos",
    "Siembra de árboles",
    "Apoyo en biblioteca municipal",
    "Asistencia a adultos mayores",
    "Campaña de reciclaje",
    "Apoyo en eventos culturales",
    "Alfabetización digital",
]


def fecha_aleatoria(inicio: datetime, fin: datetime) -> str:
    delta = (fin - inicio).days
    return (inicio + timedelta(days=random.randint(0, delta))).strftime("%Y-%m-%d")


def generar_estudiantes(n: int = 100) -> list[dict]:
    estudiantes = []
    usados = set()
    for i in range(1, n + 1):
        nombre = random.choice(NOMBRES)
        apellido = f"{random.choice(APELLIDOS)} {random.choice(APELLIDOS)}"
        estado = random.choices(
            ESTADOS_ESTUDIANTE,
            weights=[80, 8, 5, 7],
            k=1,
        )[0]
        codigo = f"BEC{2020 + random.randint(0, 4)}{i:04d}"
        fecha_ingreso = fecha_aleatoria(
            datetime(2020, 1, 1), datetime(2024, 6, 30)
        )
        est = {
            "ID": i,
            "Codigo": codigo,
            "Nombre": nombre,
            "Apellido": apellido,
            "Universidad": random.choice(UNIVERSIDADES),
            "Carrera": random.choice(CARRERAS),
            "Ciclo": str(random.randint(1, 10)),
            "Correo": f"{nombre.lower()}.{apellido.split()[0].lower()}@email.com",
            "Telefono": f"7{random.randint(100, 999)}-{random.randint(1000, 9999)}",
            "FechaIngreso": fecha_ingreso,
            "Monitor": random.choice(MONITORES),
            "Estado": estado,
            "Fotografia": "",
        }
        estudiantes.append(est)
    return estudiantes


def generar_asistencias(estudiantes: list[dict]) -> list[dict]:
    registros = []
    aid = 1
    inicio = datetime(2024, 1, 8)
    fin = datetime(2025, 6, 20)
    dias_lectivos = [
        inicio + timedelta(days=d)
        for d in range((fin - inicio).days)
        if (inicio + timedelta(days=d)).weekday() < 5  # lunes-viernes
    ]
    # Muestrear para no generar millones
    dias_muestra = random.sample(dias_lectivos, min(60, len(dias_lectivos)))
    dias_muestra.sort()

    for est in estudiantes:
        if est["Estado"] not in ("Activo", "Suspendido"):
            continue
        tendencia = random.uniform(0.55, 0.98)
        for dia in dias_muestra:
            estado = random.choices(
                ESTADOS_ASISTENCIA,
                weights=[tendencia, 1 - tendencia, 0.05, 0.03],
                k=1,
            )[0]
            registros.append({
                "ID": aid,
                "IDEstudiante": est["ID"],
                "Fecha": dia.strftime("%Y-%m-%d"),
                "Estado": estado,
                "Observacion": "" if estado == "Presente" else "Notificado al monitor",
            })
            aid += 1
    return registros


def generar_voluntariado(estudiantes: list[dict]) -> list[dict]:
    registros = []
    vid = 1
    for est in estudiantes:
        if est["Estado"] == "Egresado":
            continue
        num_actividades = random.randint(1, 8)
        for _ in range(num_actividades):
            fecha = fecha_aleatoria(datetime(2024, 1, 1), datetime(2025, 6, 1))
            registros.append({
                "ID": vid,
                "IDEstudiante": est["ID"],
                "Actividad": random.choice(ACTIVIDADES_VOLUNTARIADO),
                "Horas": round(random.uniform(2, 12), 1),
                "Fecha": fecha,
                "Observacion": random.choice(["", "Excelente participación", "Muy comprometido", ""]),
            })
            vid += 1
    return registros


def generar_seguimientos(estudiantes: list[dict]) -> list[dict]:
    registros = []
    sid = 1
    for est in estudiantes:
        if est["Estado"] == "Egresado":
            continue
        num_seg = random.randint(2, 8)
        for j in range(num_seg):
            fecha_seg = fecha_aleatoria(datetime(2024, 1, 1), datetime(2025, 5, 30))
            proximo = (
                datetime.strptime(fecha_seg, "%Y-%m-%d") + timedelta(days=random.randint(14, 45))
            ).strftime("%Y-%m-%d")
            registros.append({
                "ID": sid,
                "IDEstudiante": est["ID"],
                "Fecha": fecha_seg,
                "Tipo": random.choice(TIPOS_SEGUIMIENTO),
                "Descripcion": random.choice([
                    "Se conversó sobre avance académico y situación personal.",
                    "Reunión de seguimiento mensual.",
                    "Estudiante reporta dificultades en materia de cálculo.",
                    "Excelente avance, cumple con todos los compromisos.",
                    "Se acordó plan de mejora académica.",
                    "Seguimiento por inasistencias recientes.",
                ]),
                "AccionRealizada": random.choice([
                    "Se referió a tutor académico.",
                    "Se programó reunión con familia.",
                    "Se envió material de apoyo.",
                    "Se recordaron compromisos de beca.",
                    "Ninguna acción requerida.",
                ]),
                "ProximoSeguimiento": proximo,
                "Observaciones": "",
            })
            sid += 1
    return registros


def generar_rendimiento(estudiantes: list[dict]) -> list[dict]:
    registros = []
    rid = 1
    for est in estudiantes:
        # 2 registros por estudiante (ciclos distintos)
        for ciclo in range(1, 3):
            aprobadas = random.randint(3, 7)
            reprobadas = random.randint(0, 2)
            en_riesgo = random.randint(0, 2)
            promedio = round(random.uniform(5.5, 9.8), 2)
            fecha = fecha_aleatoria(datetime(2024, 1, 1), datetime(2025, 6, 1))
            registros.append({
                "ID": rid,
                "IDEstudiante": est["ID"],
                "Promedio": promedio,
                "MateriasAprobadas": aprobadas,
                "MateriasReprobadas": reprobadas,
                "MateriasEnRiesgo": en_riesgo,
                "FechaActualizacion": fecha,
            })
            rid += 1
    return registros


def escribir_en_excel(
    excel: ExcelManager,
    sheet_name: str,
    records: list[dict],
    columns: list[str],
) -> None:
    """Escribe registros directamente usando openpyxl para rendimiento."""
    from openpyxl import load_workbook
    from services.excel_manager import HEADER_FILL, HEADER_FONT, HEADER_ALIGN
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel.filepath)
    ws = wb[sheet_name]

    # Limpiar datos existentes (mantener encabezado)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    # Eliminar filas extra
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    THIN = Side(style="thin", color="E2E8F0")
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for row_num, record in enumerate(records, start=2):
        fill_color = "F8FAFC" if row_num % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=fill_color)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=record.get(col_name))
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(vertical="center")

    wb.save(excel.filepath)
    excel.invalidate_cache(sheet_name)


def main() -> None:
    print("🚀 Generando datos de prueba...")

    from config import (
        COLS_ESTUDIANTES, COLS_ASISTENCIAS, COLS_VOLUNTARIADO,
        COLS_SEGUIMIENTOS, COLS_RENDIMIENTO,
        SHEET_ESTUDIANTES, SHEET_ASISTENCIAS, SHEET_VOLUNTARIADO,
        SHEET_SEGUIMIENTOS, SHEET_RENDIMIENTO,
    )

    excel = ExcelManager()

    print("  ✓ Generando 100 estudiantes...")
    estudiantes = generar_estudiantes(100)
    escribir_en_excel(excel, SHEET_ESTUDIANTES, estudiantes, COLS_ESTUDIANTES)

    print("  ✓ Generando asistencias...")
    asistencias = generar_asistencias(estudiantes)
    print(f"    → {len(asistencias)} registros de asistencia")
    escribir_en_excel(excel, SHEET_ASISTENCIAS, asistencias, COLS_ASISTENCIAS)

    print("  ✓ Generando voluntariado...")
    voluntariado = generar_voluntariado(estudiantes)
    escribir_en_excel(excel, SHEET_VOLUNTARIADO, voluntariado, COLS_VOLUNTARIADO)

    print("  ✓ Generando seguimientos...")
    seguimientos = generar_seguimientos(estudiantes)
    escribir_en_excel(excel, SHEET_SEGUIMIENTOS, seguimientos, COLS_SEGUIMIENTOS)

    print("  ✓ Generando rendimiento académico...")
    rendimiento = generar_rendimiento(estudiantes)
    escribir_en_excel(excel, SHEET_RENDIMIENTO, rendimiento, COLS_RENDIMIENTO)

    print(f"\n✅ Datos generados exitosamente en: {EXCEL_FILE}")
    print(f"   Estudiantes: {len(estudiantes)}")
    print(f"   Asistencias: {len(asistencias)}")
    print(f"   Voluntariado: {len(voluntariado)}")
    print(f"   Seguimientos: {len(seguimientos)}")
    print(f"   Rendimiento: {len(rendimiento)}")


if __name__ == "__main__":
    main()
