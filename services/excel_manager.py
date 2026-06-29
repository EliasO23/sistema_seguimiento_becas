"""
ExcelManager — capa única de acceso al archivo Excel.

Toda lectura y escritura pasa por esta clase.
"""

from __future__ import annotations

import shutil
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import (
    EXCEL_FILE, BACKUPS_DIR, ALL_SHEETS,
    COLS_ESTUDIANTES, COLS_ASISTENCIAS, COLS_VOLUNTARIADO,
    COLS_SEGUIMIENTOS, COLS_RENDIMIENTO, COLS_CONFIG,
    SHEET_ESTUDIANTES, SHEET_ASISTENCIAS, SHEET_VOLUNTARIADO,
    SHEET_SEGUIMIENTOS, SHEET_RENDIMIENTO, SHEET_CONFIG,
)
from utils.logger import logger


SHEET_COLUMNS: Dict[str, List[str]] = {
    SHEET_ESTUDIANTES: COLS_ESTUDIANTES,
    SHEET_ASISTENCIAS: COLS_ASISTENCIAS,
    SHEET_VOLUNTARIADO: COLS_VOLUNTARIADO,
    SHEET_SEGUIMIENTOS: COLS_SEGUIMIENTOS,
    SHEET_RENDIMIENTO: COLS_RENDIMIENTO,
    SHEET_CONFIG: COLS_CONFIG,
}

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="E2E8F0")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class ExcelManager:
    """Gestiona toda la persistencia en el archivo Excel."""

    _lock = threading.Lock()

    def __init__(self, filepath: Path = EXCEL_FILE) -> None:
        self.filepath = filepath
        self._cache: Dict[str, pd.DataFrame] = {}
        self._cache_time: Dict[str, datetime] = {}
        self._initialize_file()

    # ── Inicialización ────────────────────────────────────────────────────────

    def _initialize_file(self) -> None:
        """Crea el archivo y las hojas si no existen."""
        if not self.filepath.exists():
            logger.info("Creando archivo Excel: %s", self.filepath)
            wb = Workbook()
            wb.remove(wb.active)  # quitar hoja vacía por defecto
            for sheet in ALL_SHEETS:
                ws = wb.create_sheet(sheet)
                self._write_header(ws, SHEET_COLUMNS[sheet])
            wb.save(self.filepath)
            logger.info("Archivo creado correctamente.")
        else:
            self._ensure_sheets()

    def _ensure_sheets(self) -> None:
        """Agrega hojas faltantes sin destruir datos existentes."""
        with self._lock:
            wb = load_workbook(self.filepath)
            changed = False
            for sheet in ALL_SHEETS:
                if sheet not in wb.sheetnames:
                    ws = wb.create_sheet(sheet)
                    self._write_header(ws, SHEET_COLUMNS[sheet])
                    changed = True
            if changed:
                wb.save(self.filepath)

    @staticmethod
    def _write_header(ws, columns: List[str]) -> None:
        """Escribe la fila de encabezados con estilo."""
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = CELL_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 6, 14)
        ws.row_dimensions[1].height = 24

    # ── Lectura ───────────────────────────────────────────────────────────────

    def read_sheet(self, sheet_name: str, use_cache: bool = True) -> pd.DataFrame:
        """Lee una hoja completa y retorna un DataFrame."""
        if use_cache and sheet_name in self._cache:
            return self._cache[sheet_name].copy()
        try:
            df = pd.read_excel(
                self.filepath,
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str,
            )
            df = df.where(pd.notna(df), other=None)
            self._cache[sheet_name] = df
            self._cache_time[sheet_name] = datetime.now()
            return df.copy()
        except Exception as exc:
            logger.error("Error leyendo hoja '%s': %s", sheet_name, exc)
            return pd.DataFrame(columns=SHEET_COLUMNS.get(sheet_name, []))

    def invalidate_cache(self, sheet_name: Optional[str] = None) -> None:
        """Invalida caché para forzar recarga desde disco."""
        if sheet_name:
            self._cache.pop(sheet_name, None)
        else:
            self._cache.clear()

    # ── Escritura ─────────────────────────────────────────────────────────────

    def _next_id(self, sheet_name: str) -> int:
        """Retorna el siguiente ID disponible para una hoja."""
        df = self.read_sheet(sheet_name)
        if df.empty or "ID" not in df.columns:
            return 1
        ids = pd.to_numeric(df["ID"], errors="coerce").dropna()
        return int(ids.max()) + 1 if not ids.empty else 1

    def insert_row(self, sheet_name: str, data: Dict[str, Any]) -> int:
        """Inserta una fila nueva y retorna el ID asignado."""
        new_id = self._next_id(sheet_name)
        data["ID"] = new_id
        with self._lock:
            wb = load_workbook(self.filepath)
            try:
                ws = wb[sheet_name]
                columns = SHEET_COLUMNS[sheet_name]
                row_values = [data.get(col) for col in columns]
                ws.append(row_values)
                self._style_data_row(ws, ws.max_row, len(columns))
                wb.save(self.filepath)
            except PermissionError as exc:
                logger.error(
                    "No se pudo guardar el archivo Excel '%s'. Probablemente está abierto en Excel u otro proceso.",
                    self.filepath,
                )
                raise RuntimeError(
                    "No se pudo guardar el registro porque el archivo Excel está en uso. Cierra Excel y vuelve a intentarlo."
                ) from exc
        self.invalidate_cache(sheet_name)
        logger.debug("Fila insertada en '%s' con ID=%s", sheet_name, new_id)
        return new_id

    def update_row(self, sheet_name: str, row_id: int, data: Dict[str, Any]) -> bool:
        """Actualiza una fila existente identificada por ID."""
        with self._lock:
            wb = load_workbook(self.filepath)
            try:
                ws = wb[sheet_name]
                columns = SHEET_COLUMNS[sheet_name]
                id_col = columns.index("ID") + 1

                for row in ws.iter_rows(min_row=2):
                    if str(row[id_col - 1].value) == str(row_id):
                        for col_idx, col_name in enumerate(columns, start=1):
                            if col_name in data and col_name != "ID":
                                row[col_idx - 1].value = data[col_name]
                        wb.save(self.filepath)
                        self.invalidate_cache(sheet_name)
                        return True
            except PermissionError as exc:
                logger.error(
                    "No se pudo actualizar el archivo Excel '%s'. Probablemente está abierto en Excel u otro proceso.",
                    self.filepath,
                )
                raise RuntimeError(
                    "No se pudo actualizar el registro porque el archivo Excel está en uso. Cierra Excel y vuelve a intentarlo."
                ) from exc
        logger.warning("ID=%s no encontrado en '%s'", row_id, sheet_name)
        return False

    def delete_row(self, sheet_name: str, row_id: int) -> bool:
        """Elimina una fila por ID."""
        with self._lock:
            wb = load_workbook(self.filepath)
            try:
                ws = wb[sheet_name]
                columns = SHEET_COLUMNS[sheet_name]
                id_col = columns.index("ID") + 1

                for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if str(row[id_col - 1].value) == str(row_id):
                        ws.delete_rows(row_num)
                        wb.save(self.filepath)
                        self.invalidate_cache(sheet_name)
                        return True
            except PermissionError as exc:
                logger.error(
                    "No se pudo eliminar en el archivo Excel '%s'. Probablemente está abierto en Excel u otro proceso.",
                    self.filepath,
                )
                raise RuntimeError(
                    "No se pudo eliminar el registro porque el archivo Excel está en uso. Cierra Excel y vuelve a intentarlo."
                ) from exc
        return False

    @staticmethod
    def _style_data_row(ws, row_num: int, num_cols: int) -> None:
        """Aplica estilo alternado a fila de datos."""
        fill_color = "F8FAFC" if row_num % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=fill_color)
        THIN_S = Side(style="thin", color="E2E8F0")
        border = Border(left=THIN_S, right=THIN_S, top=THIN_S, bottom=THIN_S)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # ── Búsqueda ──────────────────────────────────────────────────────────────

    def find_by_id(self, sheet_name: str, row_id: int) -> Optional[Dict[str, Any]]:
        """Retorna un dict con los datos de una fila o None."""
        df = self.read_sheet(sheet_name)
        if df.empty:
            return None
        df["ID"] = pd.to_numeric(df["ID"], errors="coerce")
        match = df[df["ID"] == row_id]
        if match.empty:
            return None
        return match.iloc[0].to_dict()

    def find_by_field(
        self, sheet_name: str, field: str, value: Any
    ) -> pd.DataFrame:
        """Filtra filas donde field == value."""
        df = self.read_sheet(sheet_name)
        if field not in df.columns:
            return pd.DataFrame()
        return df[df[field].astype(str) == str(value)].copy()

    def search(self, sheet_name: str, query: str) -> pd.DataFrame:
        """Búsqueda de texto libre en todas las columnas."""
        df = self.read_sheet(sheet_name)
        if df.empty:
            return df
        mask = df.apply(
            lambda col: col.astype(str).str.contains(query, case=False, na=False)
        ).any(axis=1)
        return df[mask].copy()

    # ── Backup ────────────────────────────────────────────────────────────────

    def backup(self) -> Path:
        """Crea una copia de seguridad con timestamp."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = BACKUPS_DIR / f"becados_backup_{ts}.xlsx"
        shutil.copy2(self.filepath, dest)
        # Mantener solo los 10 backups más recientes
        backups = sorted(BACKUPS_DIR.glob("becados_backup_*.xlsx"))
        for old in backups[:-10]:
            old.unlink(missing_ok=True)
        logger.info("Backup creado: %s", dest)
        return dest

    # ── Export ────────────────────────────────────────────────────────────────

    def export_sheet_csv(self, sheet_name: str, dest: Path) -> Path:
        """Exporta una hoja a CSV."""
        df = self.read_sheet(sheet_name)
        df.to_csv(dest, index=False, encoding="utf-8-sig")
        return dest

    def export_all_excel(self, dest: Path) -> Path:
        """Exporta todo el Excel a una ruta destino."""
        shutil.copy2(self.filepath, dest)
        return dest
